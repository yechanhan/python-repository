import openpyxl
import glob

xList = glob.glob('./data/*')
dList = [[] for _ in range(10)]

for x in xList:
	xfd = openpyxl.load_workbook(x, data_only=True)
	sheet = xfd['Sheet1']

	personDict = {
		'st_num' : sheet['A2'].value,
		'name' : sheet['B2'].value,
		'group_id' : sheet['C2'].value,
		'git' : sheet['D2'].value
	}
	group = int(personDict['group_id'][1:])
	dList[group-1].append(personDict)

newX = openpyxl.Workbook()
newX.remove_sheet(newX.get_sheet_by_name('Sheet'))

for i in range(10):
	newSheet = newX.create_sheet()
	newSheet.title = 'Sheet%d'%(i+1)
	newSheet.append(['st_num', 'name', 'group_id', 'git'])

	for s in dList[i]:
		newSheet.append([str(s['st_num']), s['name'], s['group_id'], s['git']])

newX.save('./group_python.xlsx')